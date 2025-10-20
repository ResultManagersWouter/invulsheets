from __future__ import annotations

from typing import Sequence
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font
# your helpers
from toelichting_invulsheet import build_intro_text
from output_sheetnames import SHEETS_OUT


def _to_excel_scalar(v):
    if v is pd.NA or v is None:
        return None
    if isinstance(v, float) and np.isnan(v):
        return None
    if isinstance(v, (list, tuple, set, dict)):
        return str(v)
    return v



def write_df(ws, df: pd.DataFrame, start_row=1, start_col=1, include_index=False):
    """Write a pandas DataFrame to a worksheet with headers (optionally include index)."""
    if df is None or df.empty:
        return (start_row, start_col, start_row, start_col)

    df2 = df.copy()
    if include_index:
        index_name = str(df2.index.name) if df2.index.name else "index"
        df2.insert(0, index_name, df2.index)

    headers = [str(h) if h is not None else "" for h in df2.columns]
    for j, col in enumerate(headers, start=start_col):
        ws.cell(row=start_row, column=j, value=col)
    for i, (_, row) in enumerate(df2.iterrows(), start=start_row + 1):
        for j, col in enumerate(df2.columns, start=start_col):
            ws.cell(row=i, column=j, value=_to_excel_scalar(row[col]))
    end_row = start_row + len(df2.index)
    end_col = start_col + len(headers) - 1
    return (start_row, start_col, end_row, end_col)


def write_single_column(ws, header: str, values: Sequence[str], start_row: int = 1, start_col: int = 1):
    """
    Write a single header + values vertically. Returns (r1, c1, r2, c2).
    """
    ws.cell(row=start_row, column=start_col, value=header)
    for i, v in enumerate(values, start=start_row + 1):
        ws.cell(row=i, column=start_col, value=_to_excel_scalar(v))
    r1 = start_row
    c1 = start_col
    r2 = start_row + len(values)  # includes header row at start_row
    c2 = start_col
    return (r1, c1, r2, c2)


def create_table(ws, start_row, start_col, end_row, end_col, display_name, style_medium=9):
    """Create an Excel Table."""
    if end_row < start_row or end_col < start_col:
        return None
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    tbl = Table(displayName=display_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=f"TableStyleMedium{style_medium}",
        showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tbl)
    return tbl
def sort_each_column_desc_na_last(df: pd.DataFrame) -> pd.DataFrame:
    """
    Sort each column independently in descending order with NaNs at the bottom.
    Works for mixed dtypes by falling back to string comparison per column.
    """
    out = df.copy()
    for col in out.columns:
        s = out[col]
        non_na = s.dropna()
        try:
            # numeric / comparable types
            sorted_vals = non_na.sort_values(ascending=False)
        except TypeError:
            # mixed / incomparable types → sort as strings
            sorted_vals = non_na.astype(str).sort_values(ascending=False)
        out[col] = list(sorted_vals) + [np.nan] * s.isna().sum()
    return out

def build_workbook_minimal(
    objecttype_tabel: pd.DataFrame | None,   # <-- renamed
    attribuuttabel: pd.DataFrame,
    domein_waarden: pd.DataFrame | None,
    columns: Sequence[str],
    output_path: str,
    fp_bomen: str,
    fp_beplanting: str,
    fp_verharding: str,
    *,
    invul_data_rows: int = 100,
) -> str:
    """
    Returns
    -------
    str
        Het `output_path`.
    """
    wb = Workbook()

    # 1️⃣ Toelichting invulsheet
    ws_toel = wb.active
    ws_toel.title = SHEETS_OUT.SHEET_TOEL.value
    intro = build_intro_text(fp_bomen, fp_beplanting, fp_verharding)
    ws_toel.merge_cells("A1:H40")
    cell = ws_toel["A1"]
    cell.value = intro
    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    cell.font = Font(size=11)
    ws_toel.column_dimensions["A"].width = 15
    for col_letter in ["B", "C", "D", "E", "F", "G", "H"]:
        ws_toel.column_dimensions[col_letter].width = 20
    ws_toel.row_dimensions[1].height = 800

    # 2️⃣ Invulsheet
    ws_invul = wb.create_sheet(SHEETS_OUT.SHEET_INVUL.value)
    invul_headers = [
        "CAD-ID", "GISIB-ID", "Bewerkingscode",
        "Objecttype", "Type", "Type gedetailleerd", "Type extra gedetailleerd"
    ]
    for j, h in enumerate(invul_headers, start=1):
        ws_invul.cell(row=1, column=j, value=h)
    last_row = 1 + invul_data_rows
    widths = [15, 18, 18, 22, 18, 26, 30]
    for idx, w in enumerate(widths, start=1):
        ws_invul.column_dimensions[get_column_letter(idx)].width = w
    create_table(ws_invul, 1, 1, last_row, len(invul_headers), "InvulTabel", style_medium=10)

    # 3️⃣ 'tabel' sheet (inclusief index)
    ws_tabel = wb.create_sheet(SHEETS_OUT.SHEET_TABEL.value)
    r1, c1, r2, c2 = write_df(ws_tabel, attribuuttabel, start_row=1, start_col=1, include_index=True)
    if r2 >= r1:
        create_table(ws_tabel, 1, 1, r2, c2, "Attribuuttabel", style_medium=9)

    # 4️⃣ 'attributen' sheet (optioneel) — per kolom sorteren (aflopend, NaN onderaan)
    if objecttype_tabel is not None and not objecttype_tabel.empty:
        df_attr = sort_each_column_desc_na_last(objecttype_tabel)
        ws_attr = wb.create_sheet(SHEETS_OUT.SHEET_ATTR.value)
        r1a, c1a, r2a, c2a = write_df(ws_attr, df_attr, start_row=1, start_col=1)
        if r2a >= r1a:
            create_table(ws_attr, 1, 1, r2a, c2a, "Taxonomie_tabel", style_medium=8)

    # 5️⃣ 'domeinwaarden' sheet (subset + fallback)
    if domein_waarden is not None and not domein_waarden.empty:
        ws_dom = wb.create_sheet(SHEETS_OUT.SHEET_DOM.value)
        cols_present = [c for c in columns if c in domein_waarden.columns] or list(domein_waarden.columns)
        domein_subset = domein_waarden.loc[:, cols_present]
        r1d, c1d, r2d, c2d = write_df(ws_dom, domein_subset, start_row=1, start_col=1)
        if r2d >= r1d and c2d >= c1d:
            create_table(ws_dom, 1, 1, r2d, c2d, "Domein_tabel", style_medium=7)

    # 6️⃣ 'variabelen' sheet (met waarden in A-kolom)
    ws_var = wb.create_sheet(SHEETS_OUT.SHEET_VAR.value)
    ws_var.column_dimensions["A"].width = 24
    bewerkingscode_values = ["Nieuw", "Verwijderen", "Aanpassen", "Instant laten"]
    r1v, c1v, r2v, c2v = write_single_column(
        ws_var, header="Bewerkingscode", values=bewerkingscode_values, start_row=1, start_col=1
    )
    create_table(ws_var, r1v, c1v, r2v, c2v, "Variabelen", style_medium=6)

    wb.save(output_path)
    return output_path
